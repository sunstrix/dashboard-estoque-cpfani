import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from datetime import datetime, timezone, timedelta
import openpyxl
from io import BytesIO
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import time
import os

# ReportLab para exportação PDF
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# ==========================================
# 1. CONFIGURAÇÃO DA PÁGINA E CSS EXPANDIDO
# ==========================================
st.set_page_config(
    page_title="Painel de Performance de Estoque NSF - CP Fani",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
/* ── Base ── */
.main { background-color: #0e1117; }
html, body, [class*="css"] { font-family: 'Segoe UI', sans-serif; }

/* ── Métricas / KPIs ── */
div[data-testid="stMetricValue"] {
    font-size: 28px; font-weight: 700;
    color: #D4AF37;
    text-shadow: 0 0 12px rgba(212,175,55,0.35);
}
div[data-testid="stMetricLabel"] { font-size: 13px; color: #8da9be; }
div[data-testid="stMetric"] {
    background: linear-gradient(135deg, #111820 0%, #0d1f14 100%);
    border: 1px solid #1a3d25;    border-left: 4px solid #007A33;
    border-radius: 10px;
    padding: 16px 20px;
    box-shadow: 0 4px 16px rgba(0,122,51,0.15);
}

/* ── Abas ── */
.stTabs [data-baseweb="tab-list"] {
    background: #0a0d12;
    border-bottom: 2px solid #007A33;
    border-radius: 6px 6px 0 0;
    gap: 4px;
    padding: 4px 8px 0;
}
.stTabs [data-baseweb="tab"] {
    color: #8da9be; font-size: 15px;
    padding: 8px 18px; border-radius: 6px 6px 0 0;
    border: none; background: transparent;
    transition: color 0.2s, background 0.2s;
}
.stTabs [data-baseweb="tab"]:hover { color: #D4AF37; background: #0d1f14; }
.stTabs [data-baseweb="tab"][aria-selected="true"] {
    color: #D4AF37; font-weight: 700;
    background: linear-gradient(180deg, #0d1f14 0%, #0a0d12 100%);
    border-top: 2px solid #007A33;
}

/* ── Botão principal ── */
div.stButton > button {
    background: linear-gradient(135deg, #007A33, #005a26);
    color: #fff; border: none; border-radius: 8px;
    padding: 0.55rem 1.2rem; font-weight: 600; letter-spacing: 0.3px;
    box-shadow: 0 2px 8px rgba(0,122,51,0.4);
    transition: all 0.2s;
}
div.stButton > button:hover {
    background: linear-gradient(135deg, #009940, #007A33);
    color: #D4AF37; box-shadow: 0 4px 14px rgba(0,122,51,0.55);
    transform: translateY(-1px);
}

/* ── Botão de download ── */
div.stDownloadButton > button {
    background: linear-gradient(135deg, #1a3d25, #0d1f14);
    color: #D4AF37; border: 1px solid #007A33; border-radius: 8px;
    padding: 0.45rem 1rem; font-weight: 600;
    transition: all 0.2s;
}
div.stDownloadButton > button:hover {
    background: linear-gradient(135deg, #007A33, #005a26);    color: #fff; border-color: #D4AF37;
}

/* ── Sidebar ── */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #080b0f 0%, #0a1510 100%);
    border-right: 2px solid #007A33;
}
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] .stSubheader { color: #D4AF37 !important; }

/* ── Selectbox ── */
div[data-baseweb="select"] > div {
    background-color: #111820 !important;
    border: 1px solid #1a3d25 !important;
    border-radius: 8px !important;
}
div[data-baseweb="select"] > div:focus-within {
    border-color: #007A33 !important;
    box-shadow: 0 0 0 2px rgba(0,122,51,0.3) !important;
}

/* ── Tabelas (dataframe) ── */
div[data-testid="stDataFrame"] {
    border: 1px solid #1a3d25;
    border-radius: 8px;
    overflow: hidden;
}

/* ── Títulos ── */
h1 { color: #D4AF37 !important; font-size: 2rem !important; letter-spacing: -0.5px; }
h2, h3 { color: #D4AF37 !important; }

/* ── Separadores ── */
hr { border-color: #1a3d25 !important; margin: 1.5rem 0; }

/* ── Spinner / info ── */
div[data-testid="stSpinner"] { color: #007A33; }

/* ── Expander ── */
details summary {
    color: #D4AF37 !important;
    background: #0d1f14;
    border: 1px solid #1a3d25;
    border-radius: 6px;
    padding: 6px 12px;
}

/* ── Barra de progresso ── */
.stProgress > div > div > div > div { background-color: #007A33; }
/* ── Scrollbar ── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #0e1117; }
::-webkit-scrollbar-thumb { background: #007A33; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #D4AF37; }
</style>
""", unsafe_allow_html=True)

# ==========================================
# CONSTANTES E MAPEAMENTOS
# ==========================================
SPREADSHEET_ID_PRINCIPAL = "1EDDyKie9UiugMLMowcPzHfViqzziFcSgxVPvZ2Rx3L0"
SPREADSHEET_ID_SEGURANCA = "1uHonFnFM4p7bz4s7YpewhKHNs6fSEfw9rDMTKC7jtHE"
SPREADSHEET_ID_DRAFT = "11Z21gFvJ9pm2xSlF3IweC7xcYZwAZWrjcWDnRe5LexY"

URL_EXCEL = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID_PRINCIPAL}/export?format=xlsx"
URL_ESTOQUE_SEGURANCA = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID_SEGURANCA}/export?format=xlsx"
URL_DRAFT = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID_DRAFT}/export?format=xlsx"

DE_PARA_LOJAS = {
    4842: "4842 - Metrópole", 5152: "5152 - Coração", 6105: "6105 - Assai Anchieta",
    6106: "6106 - Direita", 6110: "6110 - Arouche", 8001: "8001 - Dom José",
    11576: "11576 - Davó", 12055: "12055 - São Bento", 12056: "12056 - Marechal",
    12605: "12605 - Coop", 12645: "12645 - Light", 14120: "14120 - VD SBC",
    14353: "14353 - VD SP", 20371: "20371 - Luz", 21502: "21502 - Bem Barato",
    23000: "23000 - Outlet", 23379: "23379 - Assai Piraporinha"
}

# BUG 3 FIX: Dicionário reverso para extração segura de PDV
DE_PARA_LOJAS_REVERSO = {v: k for k, v in DE_PARA_LOJAS.items()}

MAPEAMENTO_PDV_DRAFT = {
    'Loja: 4842 - N. S. F. COSMETICOS E PRESENTES LTDA': 4842,
    'Loja: 5152 - N. S. F. COSMETICOS E PRESENTES LTDA': 5152,
    'Loja: 6105 - N. S. F. COSMETICOS E PRESENTES LTDA': 6105,
    'Loja: 6106 - N. S. F. COSMETICOS E PRESENTES LTDA': 6106,
    'Loja: 6110 - N. S. F. COSMETICOS E PRESENTES LTDA': 6110,
    'Loja: 8001 - N. S. F. COSMETICOS E PRESENTES LTDA': 8001,
    'Loja: 11576 - N. S. F. COSMETICOS E PRESENTES LTDA': 11576,
    'Loja: 12055 - N. S. F. COSMETICOS E PRESENTES LTDA': 12055,
    'Loja: 12056 - S. P. ARON COSMETICOS EPP': 12056,
    'Loja: 12605 - N.S.F. COSMETICOS E PRESENTES LTDA.': 12605,
    'Loja: 12645 - N. S. F. COSMETICOS E PRESENTES LTDA': 12645,
    'Loja: 14120 - ARPEL DISTRIBUIDORA DE COSMETICOS LTDA - EPP': 14120,
    'Loja: 14353 - ARPEL DISTRIBUIDORA DE COSMETICOS LTDA - EPP': 14353,
    'Loja: 20371 - N. S. F. COSMÉTICOS E PRESENTES LTDA.': 20371,
    'Loja: 21502 - N. S. F. COSMETICOS E PRESENTES LTD': 21502,
    'Loja: 23000 - N. S. F. COSMETICOS E PRESENTES LTD': 23000,
    'Loja: 23379 - N. S. F. COSMETICOS E PRESENTES LTD': 23379}

NOMES_MARCAS = {
    'BOTICARIO': 'O Boticário',
    'EUDORA': 'Eudora',
    'QUEM_DISSE_BERENICE': 'Quem Disse, Berenice?'
}

ABAS_SEGURANCA = {'BOT': 'O Boticário', 'EUD': 'Eudora', 'QDB': 'Quem Disse, Berenice?'}

LOGOS_MARCAS = {
    'O Boticário': 'logo_boticario.png',
    'Eudora': 'logo_eudora.png',
    'Quem Disse, Berenice?': 'logo_qdb.png'
}

CORES_MARCAS = {
    'O Boticário': '#007A33',
    'Eudora': '#a855f7',
    'Quem Disse, Berenice?': '#ff4b4b'
}

# ==========================================
# FUNÇÕES AUXILIARES
# ==========================================

def criar_sessao_com_retry():
    session = requests.Session()
    retry = Retry(total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504], allowed_methods=["GET"])
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session


def _registrar_erro(msg):
    """BUG 5 FIX: Registra erros silenciosos no session_state para exibição posterior."""
    if 'erros_carregamento' not in st.session_state:
        st.session_state['erros_carregamento'] = []
    st.session_state['erros_carregamento'].append(str(msg))


def download_arquivo_excel_com_retry(url, descricao="arquivo", timeout=60):
    """
    BUG 1 FIX: Lógica de fallback corrigida.
    - Sempre retorna BytesIO do primeiro request se tamanho >= esperado
    - No fallback, retorna primeiro conteúdo se segundo falhar mas primeiro tiver dados
    - Só retorna None se ambos falharem completamente
    """
    session = criar_sessao_com_retry()    primeiro_conteudo = None

    try:
        response = session.get(url, timeout=timeout, stream=True)
        response.raise_for_status()
        primeiro_conteudo = response.content

        if 'content-length' in response.headers:
            expected_size = int(response.headers['content-length'])
            actual_size = len(primeiro_conteudo)

            if actual_size >= expected_size:
                return BytesIO(primeiro_conteudo)

            # Tamanho menor que esperado → tenta novamente
            time.sleep(2)
            try:
                response2 = session.get(url, timeout=timeout, stream=True)
                response2.raise_for_status()
                segundo_conteudo = response2.content
                if len(segundo_conteudo) >= expected_size:
                    return BytesIO(segundo_conteudo)
                # Segundo também falhou → retorna primeiro se tiver dados
                if len(primeiro_conteudo) > 0:
                    return BytesIO(primeiro_conteudo)
                return None
            except Exception:
                # Segundo request falhou → retorna primeiro se tiver dados
                if len(primeiro_conteudo) > 0:
                    return BytesIO(primeiro_conteudo)
                return None
        else:
            # Sem content-length header → retorna se tiver conteúdo
            if len(primeiro_conteudo) > 0:
                return BytesIO(primeiro_conteudo)
            return None

    except Exception as e:
        _registrar_erro(f"Erro ao baixar {descricao}: {str(e)[:200]}")
        return None


def exibir_kpi_card(col, icone, titulo, valor_fmt, delta_texto=None, cor_delta="#ff4b4b"):
    """MELHORIA V2: Card de KPI customizado com ícone e badge colorido."""
    delta_html = f'<div style="font-size:12px;color:{cor_delta};margin-top:4px;">{delta_texto}</div>' if delta_texto else ''
    col.markdown(f"""
    <div style="
        background: linear-gradient(135deg,#111820,#0d1f14);
        border:1px solid #1a3d25; border-left:4px solid #007A33;
        border-radius:10px; padding:18px 20px;        box-shadow:0 4px 16px rgba(0,122,51,0.15);
        min-height:110px;
    ">
        <div style="font-size:22px;margin-bottom:4px;">{icone}</div>
        <div style="font-size:12px;color:#8da9be;margin-bottom:6px;">{titulo}</div>
        <div style="font-size:26px;font-weight:700;color:#D4AF37;
                    text-shadow:0 0 10px rgba(212,175,55,0.3);">{valor_fmt}</div>
        {delta_html}
    </div>
    """, unsafe_allow_html=True)


def exibir_titulo_marca(nome_marca, tamanho_logo=30):
    """MELHORIA V4: Seção de marca com card visual e faixa de cor."""
    cor = CORES_MARCAS.get(nome_marca, '#007A33')
    col_logo, col_nome = st.columns([0.08, 0.92])
    with col_logo:
        logo_path = LOGOS_MARCAS.get(nome_marca, '')
        if logo_path:
            try:
                st.image(logo_path, width=tamanho_logo)
            except Exception:
                st.write("🏷️")
        else:
            st.write("🏷️")
    with col_nome:
        st.markdown(f"""
        <div style="
            border-left: 4px solid {cor};
            padding: 6px 14px;
            background: linear-gradient(90deg, {cor}22, transparent);
            border-radius: 0 8px 8px 0;
            margin-bottom: 4px;
        ">
            <span style="font-size:18px; font-weight:700; color:{cor};">{nome_marca}</span>
        </div>
        """, unsafe_allow_html=True)


def obter_horario_brasilia():
    fuso_brasilia = timezone(timedelta(hours=-3))
    agora_brasilia = datetime.now(fuso_brasilia)
    return agora_brasilia.strftime("%d/%m/%Y às %H:%M:%S")


def carregar_planilha_draft(url):
    excel_buffer = download_arquivo_excel_com_retry(url, "planilha draft de custos", timeout=90)
    if excel_buffer is None:
        return pd.DataFrame()
    try:        excel_file = pd.ExcelFile(excel_buffer)
        if not excel_file.sheet_names:
            return pd.DataFrame()
        df_draft = pd.read_excel(excel_file, sheet_name=excel_file.sheet_names[0])
        if df_draft.empty:
            return pd.DataFrame()
        df_draft.columns = [str(col).strip().upper() for col in df_draft.columns]

        colunas_loja_possiveis = ['LOJA', 'PDV', 'LOJA/PDV', 'LOJA - PDV', 'CÓDIGO LOJA', 'CODIGO LOJA']
        coluna_loja = next((col for col in colunas_loja_possiveis if col in df_draft.columns), None)

        colunas_sku_possiveis = ['SKU', 'CÓDIGO', 'CODIGO', 'CÓDIGO SKU', 'CODIGO SKU', 'CÓD. SKU']
        coluna_sku = next((col for col in colunas_sku_possiveis if col in df_draft.columns), None)

        colunas_custo_possiveis = ['CUSTO', 'PREÇO DE CUSTO', 'PRECO DE CUSTO', 'CUSTO UNITÁRIO',
                                   'CUSTO UNITARIO', 'VALOR CUSTO', 'CUSTO (R$)', 'CUSTO R$']
        coluna_custo = next((col for col in colunas_custo_possiveis if col in df_draft.columns), None)

        if coluna_custo is None and len(df_draft.columns) > 9:
            coluna_custo = df_draft.columns[9]

        if coluna_loja is None or coluna_sku is None or coluna_custo is None:
            _registrar_erro(f"Planilha Draft: colunas faltantes. Loja={coluna_loja}, SKU={coluna_sku}, Custo={coluna_custo}")
            return pd.DataFrame()

        df_resultado = pd.DataFrame()
        df_resultado['LOJA_NOME'] = df_draft[coluna_loja].astype(str).str.strip()
        df_resultado['SKU'] = df_draft[coluna_sku].astype(str).str.strip()
        df_resultado['CUSTO_DRAFT'] = pd.to_numeric(df_draft[coluna_custo], errors='coerce').fillna(0)
        df_resultado['PDV'] = df_resultado['LOJA_NOME'].map(MAPEAMENTO_PDV_DRAFT)
        df_resultado = df_resultado[df_resultado['PDV'].notna()].copy()
        df_resultado['PDV'] = df_resultado['PDV'].astype(int)
        return df_resultado[['PDV', 'SKU', 'CUSTO_DRAFT']].copy()
    except Exception as e:
        _registrar_erro(f"Erro ao processar planilha draft: {str(e)[:200]}")
        return pd.DataFrame()


def carregar_estoque_seguranca(url):
    excel_buffer = download_arquivo_excel_com_retry(url, "planilha de estoque de segurança", timeout=90)
    if excel_buffer is None:
        return pd.DataFrame()
    try:
        excel_file = pd.ExcelFile(excel_buffer)
        abas_esperadas = ['BOT', 'EUD', 'QDB']
        abas_encontradas = [aba for aba in abas_esperadas if aba.upper() in [a.upper() for a in excel_file.sheet_names]]
        if not abas_encontradas:
            _registrar_erro(f"Planilha Segurança: nenhuma das abas esperadas encontrada. Disponíveis: {excel_file.sheet_names}")
            return pd.DataFrame()
        dfs_abas = []
        for aba_nome in abas_encontradas:
            try:
                aba_exata = [nome for nome in excel_file.sheet_names if nome.upper() == aba_nome.upper()][0]
                df_abas = pd.read_excel(excel_file, sheet_name=aba_exata)
                if df_abas.empty:
                    continue
                df_abas.columns = [col.strip().upper() for col in df_abas.columns]
                if 'PDV' not in df_abas.columns or 'SKU' not in df_abas.columns:
                    continue
                colunas_possiveis = ['ESTOQUE DE SEGURANCA', 'ESTOQUE_DE_SEGURANCA', 'ESTOQUE_SEGURANCA',
                                   'ESTOQUE MINIMO', 'ESTOQUE_MINIMO', 'MINIMO', 'SEGURANCA', 'QTD_MINIMA']
                coluna_seguranca = next((col for col in colunas_possiveis if col in df_abas.columns), None)
                if coluna_seguranca is None:
                    df_abas['ESTOQUE_DE_SEGURANCA'] = 0
                else:
                    df_abas = df_abas.rename(columns={coluna_seguranca: 'ESTOQUE_DE_SEGURANCA'})
                    df_abas['ESTOQUE_DE_SEGURANCA'] = pd.to_numeric(df_abas['ESTOQUE_DE_SEGURANCA'], errors='coerce').fillna(0)
                df_abas['PDV'] = pd.to_numeric(df_abas['PDV'], errors='coerce')
                df_abas['SKU'] = df_abas['SKU'].astype(str).str.strip()
                df_abas['MARCA_REFERENCIA'] = ABAS_SEGURANCA.get(aba_nome, aba_nome)
                dfs_abas.append(df_abas[['PDV', 'SKU', 'ESTOQUE_DE_SEGURANCA', 'MARCA_REFERENCIA']].copy())
            except Exception as e:
                _registrar_erro(f"Erro ao processar aba {aba_nome} da segurança: {str(e)[:150]}")
                continue

        if dfs_abas:
            return pd.concat(dfs_abas, ignore_index=True)
        return pd.DataFrame()
    except Exception as e:
        _registrar_erro(f"Erro ao carregar planilha de segurança: {str(e)[:200]}")
        return pd.DataFrame()


def obter_data_atualizacao_planilha(url_excel):
    excel_buffer = download_arquivo_excel_com_retry(url_excel, "metadados da planilha", timeout=60)
    if excel_buffer is None:
        return None
    try:
        workbook = openpyxl.load_workbook(excel_buffer, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=1, max_row=10, max_col=5):
                for cell in row:
                    if cell.value:
                        valor_str = str(cell.value).strip().upper()
                        if any(p in valor_str for p in ['/', ':', '202', '203']):
                            if isinstance(cell.value, datetime):
                                workbook.close()
                                return cell.value.strftime("%d/%m/%Y às %H:%M:%S")                            elif isinstance(cell.value, str):
                                workbook.close()
                                return cell.value
        workbook.close()
        return None
    except Exception as e:
        _registrar_erro(f"Erro ao ler metadados: {str(e)[:150]}")
        return None


@st.cache_data(ttl=3600)
def carregar_dados_nuvem(url_principal, url_seguranca, url_draft):
    dicionario_marcas = {}
    data_atualizacao = None
    try:
        df_estoque_seguranca = carregar_estoque_seguranca(url_seguranca)
        df_draft = carregar_planilha_draft(url_draft)
        data_atualizacao = obter_data_atualizacao_planilha(url_principal)

        excel_buffer = download_arquivo_excel_com_retry(url_principal, "planilha principal de estoque", timeout=120)
        if excel_buffer is None:
            _registrar_erro("Não foi possível baixar a planilha principal.")
            return {}, data_atualizacao

        try:
            excel_file = pd.ExcelFile(excel_buffer)
            for aba_excel, nome_exibicao in NOMES_MARCAS.items():
                if aba_excel not in excel_file.sheet_names:
                    _registrar_erro(f"Aba '{aba_excel}' não encontrada na planilha principal.")
                    continue

                df = pd.read_excel(excel_file, sheet_name=aba_excel)
                df['Marca'] = nome_exibicao
                df['PDV'] = pd.to_numeric(df['PDV'], errors='coerce')
                df['Estoque Atual'] = pd.to_numeric(df['Estoque Atual'], errors='coerce').fillna(0)
                df['Preço tabela'] = pd.to_numeric(df['Preço tabela'], errors='coerce').fillna(0)
                df['SKU'] = df['SKU'].astype(str).str.strip()

                # BUG 2 FIX: Vetorização do cálculo de custo final
                df['CUSTO_DRAFT'] = 0.0
                if not df_draft.empty:
                    df_draft_merge = df_draft[['PDV', 'SKU', 'CUSTO_DRAFT']].copy()
                    df = df.merge(df_draft_merge, on=['PDV', 'SKU'], how='left')
                    df['CUSTO_DRAFT'] = df['CUSTO_DRAFT'].fillna(0)

                df['Preço de Custo'] = np.where(
                    (df['Preço tabela'] == 0) | (df['Preço tabela'].isna()),
                    df['CUSTO_DRAFT'],
                    np.where(
                        df['CUSTO_DRAFT'] > 0,                        np.maximum(df['Preço tabela'], df['CUSTO_DRAFT']),
                        df['Preço tabela']
                    )
                )

                if 'CUSTO_DRAFT' in df.columns:
                    df = df.drop(columns=['CUSTO_DRAFT'])

                # Merge com estoque de segurança
                if not df_estoque_seguranca.empty:
                    df_seguranca_marca = df_estoque_seguranca[
                        df_estoque_seguranca['MARCA_REFERENCIA'] == nome_exibicao
                    ].copy()
                    if not df_seguranca_marca.empty:
                        df = df.merge(df_seguranca_marca[['PDV', 'SKU', 'ESTOQUE_DE_SEGURANCA']],
                                      on=['PDV', 'SKU'], how='left')
                        df['Estoque_Minimo_Qtd'] = df['ESTOQUE_DE_SEGURANCA'].fillna(0)
                        df = df.drop(columns=['ESTOQUE_DE_SEGURANCA'])
                    else:
                        regras_minimo = {'A': 15, 'B': 10, 'C': 5, 'E': 2}
                        df['Estoque_Minimo_Qtd'] = df['Classe'].map(regras_minimo).fillna(2)
                else:
                    regras_minimo = {'A': 15, 'B': 10, 'C': 5, 'E': 2}
                    df['Estoque_Minimo_Qtd'] = df['Classe'].map(regras_minimo).fillna(2)

                df['Valor_Estoque_Atual'] = df['Estoque Atual'] * df['Preço tabela']
                df['Valor_Estoque_Minimo'] = df['Estoque_Minimo_Qtd'] * df['Preço tabela']
                df['Qtd_Excesso'] = (df['Estoque Atual'] - df['Estoque_Minimo_Qtd']).clip(lower=0)
                df['Valor_Excesso'] = df['Qtd_Excesso'] * df['Preço tabela']
                df['Qtd_Falta'] = (df['Estoque_Minimo_Qtd'] - df['Estoque Atual']).clip(lower=0)
                df['Valor_Falta'] = df['Qtd_Falta'] * df['Preço tabela']
                df['Valor_Custo_Estoque_Atual'] = df['Estoque Atual'] * df['Preço de Custo']
                df['Valor_Custo_Estoque_Minimo'] = df['Estoque_Minimo_Qtd'] * df['Preço de Custo']

                dicionario_marcas[nome_exibicao] = df

        except Exception as e:
            _registrar_erro(f"Erro ao processar planilha principal: {str(e)[:200]}")
            return {}, data_atualizacao

    except Exception as e:
        _registrar_erro(f"Erro geral no carregamento: {str(e)[:200]}")

    return dicionario_marcas, data_atualizacao


def gerar_pdf_dashboard(dados_filtrados, pdv_selecionado, loja_selecionada_nome,
                        marca_selecionada, horario_brasilia,
                        v_estoque_atual_total, v_estoque_min_total,
                        v_excesso_total_total, v_falta_total_total, qtd_itens_total):    """NOVA FUNCIONALIDADE: Exportar dashboard para PDF via ReportLab."""
    if not REPORTLAB_AVAILABLE:
        st.error("Biblioteca reportlab não instalada. Adicione 'reportlab' ao requirements.txt.")
        return None

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    COR_VERDE = colors.HexColor('#007A33')
    COR_DOURADO = colors.HexColor('#D4AF37')
    COR_TEXTO = colors.HexColor('#333333')
    COR_AMARELO = colors.HexColor('#b45309')
    COR_VERMELHO = colors.HexColor('#dc2626')

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle('titulo', fontSize=16, textColor=COR_VERDE, fontName='Helvetica-Bold', spaceAfter=4)
    estilo_subtitulo = ParagraphStyle('subtitulo', fontSize=11, textColor=COR_VERDE, fontName='Helvetica-Bold', spaceAfter=4)
    estilo_normal = ParagraphStyle('normal', fontSize=9, textColor=COR_TEXTO, fontName='Helvetica', spaceAfter=2)
    estilo_rodape = ParagraphStyle('rodape', fontSize=8, textColor=colors.HexColor('#6b7e8a'), fontName='Helvetica', alignment=1)

    elementos = []

    # Logo + Cabeçalho
    try:
        if os.path.exists('logo_cp_fani.png'):
            logo = RLImage('logo_cp_fani.png', width=3*cm, height=2*cm)
            elementos.append(logo)
    except Exception:
        pass

    elementos.append(Paragraph("Painel de Performance de Estoque NSF · CP Fani", estilo_titulo))
    elementos.append(Paragraph(f"PDV: {loja_selecionada_nome}  |  Marca: {marca_selecionada}", estilo_subtitulo))
    elementos.append(Paragraph(f"Gerado em: {horario_brasilia}", estilo_normal))
    elementos.append(Spacer(1, 0.5*cm))

    # Tabela de KPIs
    dados_kpi = [
        ['KPI', 'Valor'],
        ['Valor em Estoque (Tabela)', f"R$ {v_estoque_atual_total:,.2f}"],
        ['Estoque Mínimo (Tabela)', f"R$ {v_estoque_min_total:,.2f}"],
        ['Capital Preso (Excesso)', f"R$ {v_excesso_total_total:,.2f}"],
        ['Risco de Ruptura (Falta)', f"R$ {v_falta_total_total:,.2f}"],
        ['Total de Unidades', f"{int(qtd_itens_total):,}"],
    ]
    tabela_kpi = Table(dados_kpi, colWidths=[10*cm, 6*cm])
    tabela_kpi.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), COR_VERDE),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8faf9')),
        ('TEXTCOLOR', (0, 1), (0, -1), COR_TEXTO),
        ('TEXTCOLOR', (1, 1), (1, -1), COR_VERDE),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8faf9'), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#1a3d25')),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elementos.append(tabela_kpi)
    elementos.append(Spacer(1, 0.5*cm))

    # Tabelas de Excessos e Faltas por marca
    for nome_marca, df_completo in dados_filtrados.items():
        df_loja = df_completo[df_completo['PDV'] == pdv_selecionado]
        if df_loja.empty:
            continue

        # Excessos
        df_exc = df_loja[
            (df_loja['Valor_Excesso'] > 0) & (df_loja['Estoque_Minimo_Qtd'] > 0)
        ].sort_values('Valor_Excesso', ascending=False).head(20)

        if not df_exc.empty:
            elementos.append(Paragraph(f"Excessos Críticos — {nome_marca}", estilo_subtitulo))
            colunas_exc = ['SKU', 'Descrição', 'Classe', 'Estoque Atual', 'Estoque_Minimo_Qtd', 'Qtd_Excesso', 'Valor_Excesso']
            colunas_exc_existentes = [c for c in colunas_exc if c in df_exc.columns]
            dados_exc = [colunas_exc_existentes]
            for _, row in df_exc[colunas_exc_existentes].iterrows():
                linha = []
                for col in colunas_exc_existentes:
                    v = row[col]
                    if col == 'Valor_Excesso':
                        linha.append(f"R$ {float(v):,.2f}")
                    else:
                        linha.append(str(v))
                dados_exc.append(linha)
            t_exc = Table(dados_exc, repeatRows=1)
            t_exc.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), COR_VERDE),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8faf9')),
                ('TEXTCOLOR', (0, 1), (-1, -1), COR_TEXTO),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8faf9'), colors.white]),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#1a3d25')),
                ('PADDING', (0, 0), (-1, -1), 4),
                ('TEXTCOLOR', (-1, 1), (-1, -1), COR_AMARELO),            ]))
            elementos.append(t_exc)
            elementos.append(Spacer(1, 0.4*cm))

        # Faltas
        df_flt = df_loja[df_loja['Valor_Falta'] > 0].sort_values('Valor_Falta', ascending=False).head(20)

        if not df_flt.empty:
            elementos.append(Paragraph(f"Produtos em Falta / Ruptura — {nome_marca}", estilo_subtitulo))
            colunas_flt = ['SKU', 'Descrição', 'Classe', 'Estoque Atual', 'Estoque_Minimo_Qtd', 'Qtd_Falta', 'Valor_Falta']
            colunas_flt_existentes = [c for c in colunas_flt if c in df_flt.columns]
            dados_flt = [colunas_flt_existentes]
            for _, row in df_flt[colunas_flt_existentes].iterrows():
                linha = []
                for col in colunas_flt_existentes:
                    v = row[col]
                    if col == 'Valor_Falta':
                        linha.append(f"R$ {float(v):,.2f}")
                    else:
                        linha.append(str(v))
                dados_flt.append(linha)
            t_flt = Table(dados_flt, repeatRows=1)
            t_flt.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), COR_VERMELHO),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fef2f2')),
                ('TEXTCOLOR', (0, 1), (-1, -1), COR_TEXTO),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#fef2f2'), colors.white]),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#fca5a5')),
                ('PADDING', (0, 0), (-1, -1), 4),
                ('TEXTCOLOR', (-1, 1), (-1, -1), COR_VERMELHO),
            ]))
            elementos.append(t_flt)
            elementos.append(Spacer(1, 0.4*cm))

    # Rodapé
    elementos.append(Spacer(1, 1*cm))
    elementos.append(Paragraph(f"Grupo NSF · CP Fani  |  Gerado em: {horario_brasilia}", estilo_rodape))

    try:
        doc.build(elementos)
        buffer.seek(0)
        return buffer
    except Exception as e:
        _registrar_erro(f"Erro ao gerar PDF: {str(e)[:200]}")
        return None

# ==========================================
# CARREGAMENTO DE DADOS
# ==========================================
with st.spinner("Carregando dados..."):
    dados_marcas, data_atualizacao_planilha = carregar_dados_nuvem(URL_EXCEL, URL_ESTOQUE_SEGURANCA, URL_DRAFT)
    horario_carregamento = obter_horario_brasilia()

# BUG 5 FIX: Exibir avisos de carregamento acumulados
if st.session_state.get('erros_carregamento'):
    with st.expander("⚠️ Avisos de carregamento (clique para ver)", expanded=False):
        for erro in st.session_state['erros_carregamento']:
            st.warning(erro)

if not dados_marcas:
    st.error("❌ Nenhum dado foi carregado. Verifique as permissões de compartilhamento da planilha e sua conexão com a internet.")
    st.stop()

if data_atualizacao_planilha:
    horario_exibicao = data_atualizacao_planilha
    info_timestamp = "🕒 Última atualização da planilha"
else:
    horario_exibicao = horario_carregamento
    info_timestamp = "🕒 Horário de carregamento do dashboard"

# ==========================================
# MELHORIA V3: CABEÇALHO INSTITUCIONAL APRIMORADO
# ==========================================
col_logo, col_info = st.columns([1, 3])

with col_logo:
    try:
        st.image("logo_cp_fani.png", width=180)
    except Exception:
        pass

with col_info:
    st.markdown(f"""
    <div style="
        background: linear-gradient(135deg, #0d1f14 0%, #080b0f 100%);
        border: 1px solid #1a3d25; border-left: 5px solid #007A33;
        border-radius: 12px; padding: 20px 28px; margin-bottom: 8px;
        display: flex; align-items: center; gap: 20px;
    ">
        <div>
            <div style="font-size:13px; color:#8da9be; margin-bottom:4px; letter-spacing:1px; text-transform:uppercase;">
                Grupo NSF · CP Fani
            </div>
            <div style="font-size:26px; font-weight:700; color:#D4AF37; line-height:1.2;">
                📊 Painel de Controle de Estoques e Ruptura
            </div>            <div style="font-size:12px; color:#6b7e8a; margin-top:6px;">
                {info_timestamp}: <span style="color:#a3b8cc;">{horario_exibicao}</span> · Horário de Brasília · Fonte: Google Sheets
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

st.markdown("---")

# ==========================================
# SIDEBAR E FILTROS
# ==========================================
st.sidebar.title("Filtros de Visualização")

# BUG 4 FIX: Consolidação de PDVs de todas as marcas
todos_pdvs = sorted(set(
    int(pdv)
    for df in dados_marcas.values()
    for pdv in df['PDV'].dropna()
))
opcoes_selectbox = [DE_PARA_LOJAS.get(pdv, f"PDV {pdv}") for pdv in todos_pdvs]

loja_selecionada_nome = st.sidebar.selectbox("Selecione a Loja / PDV:", opcoes_selectbox)

# BUG 3 FIX: Extração segura de PDV via dicionário reverso
pdv_selecionado = DE_PARA_LOJAS_REVERSO.get(loja_selecionada_nome)
if pdv_selecionado is None:
    st.error("PDV não reconhecido. Por favor, selecione novamente.")
    st.stop()

st.sidebar.markdown("---")

st.sidebar.subheader("Filtro de Marca")
opcoes_marca = ["Todas as Marcas"] + list(dados_marcas.keys())
marca_selecionada = st.sidebar.selectbox("Selecione a Marca:", opcoes_marca)

st.sidebar.markdown("**Marcas disponíveis:**")
col_logos_sidebar = st.sidebar.columns(len(dados_marcas))
for idx, (nome_marca, df_marca) in enumerate(dados_marcas.items()):
    with col_logos_sidebar[idx]:
        logo_path = LOGOS_MARCAS.get(nome_marca, '')
        if logo_path:
            try:
                st.image(logo_path, width=40)
            except Exception:
                pass
        st.caption(nome_marca)

st.sidebar.markdown("---")
if st.sidebar.button("🔄 Forçar Atualização dos Dados"):    st.cache_data.clear()
    st.rerun()

# MELHORIA V6: Subtítulo da loja como badge visual
st.markdown(f"""
<div style="
    display:inline-block;
    background: linear-gradient(135deg,#0d1f14,#111820);
    border:1px solid #007A33; border-radius:20px;
    padding:6px 20px; margin-bottom:12px;
">
    <span style="color:#8da9be;font-size:13px;">🏪 PDV selecionado: </span>
    <span style="color:#D4AF37;font-weight:700;font-size:15px;">{loja_selecionada_nome}</span>
</div>
""", unsafe_allow_html=True)

# Filtra dados pela marca selecionada
if marca_selecionada == "Todas as Marcas":
    dados_filtrados = dados_marcas
    titulo_secao = "Consolidado Geral (Todas as Marcas)"
else:
    dados_filtrados = {marca_selecionada: dados_marcas[marca_selecionada]}
    titulo_secao = f"Análise: {marca_selecionada}"

st.markdown(f"### {titulo_secao}")
st.markdown("---")

# ==========================================
# KPIs CONSOLIDADOS (MELHORIA V2)
# ==========================================
v_estoque_atual_total = 0
v_estoque_min_total = 0
v_excesso_total_total = 0
v_falta_total_total = 0
qtd_itens_total = 0

for nome_marca, df_completo in dados_filtrados.items():
    df_loja = df_completo[df_completo['PDV'] == pdv_selecionado]
    if not df_loja.empty:
        v_estoque_atual_total += df_loja['Valor_Estoque_Atual'].sum()
        v_estoque_min_total += df_loja['Valor_Estoque_Minimo'].sum()
        v_excesso_total_total += df_loja['Valor_Excesso'].sum()
        v_falta_total_total += df_loja['Valor_Falta'].sum()
        qtd_itens_total += df_loja['Estoque Atual'].sum()

col1, col2, col3, col4 = st.columns(4)

exibir_kpi_card(col1, "💰", "Valor em Estoque (Tabela)", f"R$ {v_estoque_atual_total:,.2f}")
exibir_kpi_card(col2, "📉", "Estoque Mínimo (Tabela)", f"R$ {v_estoque_min_total:,.2f}")
pct_excesso = f"{((v_excesso_total_total/v_estoque_atual_total)*100 if v_estoque_atual_total > 0 else 0):.1f}% do estoque"
exibir_kpi_card(col3, "⚠️", "Capital Preso (Excesso)", f"R$ {v_excesso_total_total:,.2f}", delta_texto=pct_excesso, cor_delta="#f59e0b")

exibir_kpi_card(col4, "🚨", "Risco de Ruptura (Falta)", f"R$ {v_falta_total_total:,.2f}", delta_texto="Abaixo do Mínimo", cor_delta="#ef4444")

st.markdown("---")

# ==========================================
# GRÁFICO COMPARATIVO POR MARCA
# ==========================================
if marca_selecionada == "Todas as Marcas":
    st.subheader("📊 Comparativo entre Marcas")

    dados_grafico = []
    for nome_marca, df_completo in dados_marcas.items():
        df_loja = df_completo[df_completo['PDV'] == pdv_selecionado]
        if not df_loja.empty:
            dados_grafico.append({
                'Marca': nome_marca,
                'Qtd Itens': int(df_loja['Estoque Atual'].sum()),
                'Custo Total': df_loja['Valor_Custo_Estoque_Atual'].sum()
            })

    if dados_grafico:
        df_grafico_marcas = pd.DataFrame(dados_grafico)
        col_graf1, col_graf2 = st.columns(2)

        with col_graf1:
            fig_qtd = go.Figure()
            fig_qtd.add_trace(go.Bar(
                x=df_grafico_marcas['Marca'], y=df_grafico_marcas['Qtd Itens'],
                name='Quantidade de Itens',
                marker_color=[CORES_MARCAS.get(m, '#007A33') for m in df_grafico_marcas['Marca']],
                text=[f"{v:,.0f}" for v in df_grafico_marcas['Qtd Itens']], textposition='auto'
            ))
            fig_qtd.update_layout(template='plotly_dark', plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                                  height=400, title='Quantidade de Itens por Marca', yaxis_title='Qtd de Unidades')
            st.plotly_chart(fig_qtd, use_container_width=True)

        with col_graf2:
            fig_custo = go.Figure()
            fig_custo.add_trace(go.Bar(
                x=df_grafico_marcas['Marca'], y=df_grafico_marcas['Custo Total'],
                name='Custo Total',
                marker_color=[CORES_MARCAS.get(m, '#007A33') for m in df_grafico_marcas['Marca']],
                text=[f"R$ {v:,.0f}" for v in df_grafico_marcas['Custo Total']], textposition='auto'
            ))
            fig_custo.update_layout(template='plotly_dark', plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                                    height=400, title='Custo Total por Marca (Draft + Tabela)', yaxis_title='Valor (R$)')
            st.plotly_chart(fig_custo, use_container_width=True)
# ==========================================
# CUSTO POR CURVA
# ==========================================
st.markdown("---")
st.subheader("📊 Custo Total por Curva de Produto")

df_curva_consolidado = pd.DataFrame()
for nome_marca, df_completo in dados_filtrados.items():
    df_loja = df_completo[df_completo['PDV'] == pdv_selecionado]
    if not df_loja.empty and 'Classe' in df_loja.columns:
        df_agrupado = df_loja.groupby('Classe').agg({'Valor_Custo_Estoque_Atual': 'sum', 'SKU': 'count'}).reset_index()
        df_agrupado.columns = ['Curva', 'Custo Total', 'Qtd SKUs']
        df_agrupado['Marca'] = nome_marca
        df_curva_consolidado = pd.concat([df_curva_consolidado, df_agrupado], ignore_index=True)

if not df_curva_consolidado.empty:
    if marca_selecionada == "Todas as Marcas":
        df_pivot = df_curva_consolidado.pivot_table(values='Custo Total', index='Curva', columns='Marca', aggfunc='sum', fill_value=0).reset_index()
        colunas_marcas = [col for col in df_pivot.columns if col != 'Curva']
        df_pivot['Total Geral'] = df_pivot[colunas_marcas].sum(axis=1)
        linha_total = {'Curva': 'TOTAL'}
        for col in colunas_marcas:
            linha_total[col] = df_pivot[col].sum()
        linha_total['Total Geral'] = df_pivot['Total Geral'].sum()
        df_pivot = pd.concat([df_pivot, pd.DataFrame([linha_total])], ignore_index=True)
        for col in colunas_marcas + ['Total Geral']:
            df_pivot[col] = df_pivot[col].apply(lambda x: f"R$ {x:,.2f}")
        st.dataframe(df_pivot, use_container_width=True, hide_index=True)
    else:
        df_exibicao = df_curva_consolidado[['Curva', 'Custo Total', 'Qtd SKUs']].copy().sort_values('Curva')
        total_custo = df_exibicao['Custo Total'].sum()
        total_skus = df_exibicao['Qtd SKUs'].sum()
        df_total = pd.DataFrame([{'Curva': 'TOTAL', 'Custo Total': total_custo, 'Qtd SKUs': total_skus}])
        df_exibicao = pd.concat([df_exibicao, df_total], ignore_index=True)
        df_exibicao['Custo Total'] = df_exibicao['Custo Total'].apply(lambda x: f"R$ {x:,.2f}")
        st.dataframe(df_exibicao, use_container_width=True, hide_index=True)

    fig_custo = go.Figure()
    if marca_selecionada == "Todas as Marcas":
        for nome_marca in dados_filtrados.keys():
            df_mc = df_curva_consolidado[df_curva_consolidado['Marca'] == nome_marca]
            if not df_mc.empty:
                fig_custo.add_trace(go.Bar(x=df_mc['Curva'], y=df_mc['Custo Total'], name=nome_marca,
                                           marker_color=CORES_MARCAS.get(nome_marca, '#007A33')))
        fig_custo.update_layout(barmode='stack')
    else:
        fig_custo.add_trace(go.Bar(x=df_curva_consolidado['Curva'], y=df_curva_consolidado['Custo Total'],
                                   marker_color=[CORES_MARCAS.get(marca_selecionada, '#007A33')] * len(df_curva_consolidado),
                                   text=[f"R$ {v:,.2f}" for v in df_curva_consolidado['Custo Total']], textposition='auto'))    fig_custo.update_layout(template='plotly_dark', plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                            height=400, xaxis_title='Curva', yaxis_title='Custo Total (R$)', title='Distribuição de Custo por Curva')
    st.plotly_chart(fig_custo, use_container_width=True)

# ==========================================
# ANÁLISE POR CATEGORIA
# ==========================================
st.markdown("---")
st.subheader("📊 Análise por Categoria")

df_categoria_consolidado = pd.DataFrame()
for nome_marca, df_completo in dados_filtrados.items():
    df_loja = df_completo[df_completo['PDV'] == pdv_selecionado]
    if not df_loja.empty and 'Categoria' in df_loja.columns:
        df_cat = df_loja.groupby('Categoria')[['Valor_Estoque_Atual', 'Valor_Estoque_Minimo']].sum().reset_index()
        df_cat['Marca'] = nome_marca
        df_categoria_consolidado = pd.concat([df_categoria_consolidado, df_cat], ignore_index=True)

if not df_categoria_consolidado.empty:
    fig_cat = go.Figure()
    if marca_selecionada == "Todas as Marcas":
        for nome_marca in dados_filtrados.keys():
            df_mc = df_categoria_consolidado[df_categoria_consolidado['Marca'] == nome_marca]
            if not df_mc.empty:
                fig_cat.add_trace(go.Bar(x=df_mc['Categoria'], y=df_mc['Valor_Estoque_Atual'], name=nome_marca,
                                         marker_color=CORES_MARCAS.get(nome_marca, '#007A33')))
        fig_cat.update_layout(barmode='group')
    else:
        fig_cat.add_trace(go.Bar(x=df_categoria_consolidado['Categoria'], y=df_categoria_consolidado['Valor_Estoque_Atual'],
                                 name='Estoque Atual', marker_color=CORES_MARCAS.get(marca_selecionada, '#007A33')))
        fig_cat.add_trace(go.Bar(x=df_categoria_consolidado['Categoria'], y=df_categoria_consolidado['Valor_Estoque_Minimo'],
                                 name='Estoque Mínimo', marker_color='#ff4b4b'))
        fig_cat.update_layout(barmode='group')
    fig_cat.update_layout(template='plotly_dark', plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                          height=400, xaxis_title='Categoria', yaxis_title='Valor (R$)', title='Estoque por Categoria')
    st.plotly_chart(fig_cat, use_container_width=True)

# ==========================================
# TABELAS DE EXCESSOS E FALTAS (MELHORIA V5)
# ==========================================
st.markdown("---")

for nome_marca, df_completo in dados_filtrados.items():
    df_loja = df_completo[df_completo['PDV'] == pdv_selecionado]
    if df_loja.empty:
        continue

    exibir_titulo_marca(nome_marca, tamanho_logo=35)

    col_tab1, col_tab2 = st.columns(2)
    with col_tab1:
        st.write("### 🛑 Excessos Críticos")
        df_excesso_tabela = df_loja[
            (df_loja['Valor_Excesso'] > 0) & (df_loja['Estoque_Minimo_Qtd'] > 0)
        ][['SKU', 'Descrição', 'Classe', 'Estoque Atual', 'Estoque_Minimo_Qtd', 'Qtd_Excesso', 'Preço tabela', 'Preço de Custo', 'Valor_Excesso']
        ].sort_values(by='Valor_Excesso', ascending=False)

        # MELHORIA V5: Mini-indicadores de excesso
        total_excesso_qtd = int(df_excesso_tabela['Qtd_Excesso'].sum()) if not df_excesso_tabela.empty else 0
        total_excesso_val = df_excesso_tabela['Valor_Excesso'].sum() if not df_excesso_tabela.empty else 0
        skus_excesso = len(df_excesso_tabela)
        col_e1, col_e2, col_e3 = st.columns(3)
        col_e1.metric("SKUs com excesso", skus_excesso)
        col_e2.metric("Unidades em excesso", f"{total_excesso_qtd:,}")
        col_e3.metric("Valor total em excesso", f"R$ {total_excesso_val:,.2f}")

        st.dataframe(df_excesso_tabela.style.format({'Preço tabela': 'R$ {:.2f}', 'Preço de Custo': 'R$ {:.2f}', 'Valor_Excesso': 'R$ {:.2f}'}),
                     use_container_width=True, height=280)

    with col_tab2:
        st.write("### 🚨 Produtos Críticos em Falta / Ruptura")
        df_falta_tabela = df_loja[df_loja['Valor_Falta'] > 0][
            ['SKU', 'Descrição', 'Classe', 'Estoque Atual', 'Estoque_Minimo_Qtd', 'Qtd_Falta', 'Preço tabela', 'Preço de Custo', 'Valor_Falta']
        ].sort_values(by='Valor_Falta', ascending=False)

        # MELHORIA V5: Mini-indicadores de falta
        total_falta_qtd = int(df_falta_tabela['Qtd_Falta'].sum()) if not df_falta_tabela.empty else 0
        total_falta_val = df_falta_tabela['Valor_Falta'].sum() if not df_falta_tabela.empty else 0
        skus_falta = len(df_falta_tabela)
        col_f1, col_f2, col_f3 = st.columns(3)
        col_f1.metric("SKUs em falta", skus_falta)
        col_f2.metric("Unidades em falta", f"{total_falta_qtd:,}")
        col_f3.metric("Risco financeiro", f"R$ {total_falta_val:,.2f}")

        st.dataframe(df_falta_tabela.style.format({'Preço tabela': 'R$ {:.2f}', 'Preço de Custo': 'R$ {:.2f}', 'Valor_Falta': 'R$ {:.2f}'}),
                     use_container_width=True, height=280)

    st.markdown("---")

# ==========================================
# EXPORTAR PARA PDF
# ==========================================
st.markdown("---")
pdf_buffer = gerar_pdf_dashboard(
    dados_filtrados=dados_filtrados,
    pdv_selecionado=pdv_selecionado,
    loja_selecionada_nome=loja_selecionada_nome,
    marca_selecionada=marca_selecionada,
    horario_brasilia=obter_horario_brasilia(),    v_estoque_atual_total=v_estoque_atual_total,
    v_estoque_min_total=v_estoque_min_total,
    v_excesso_total_total=v_excesso_total_total,
    v_falta_total_total=v_falta_total_total,
    qtd_itens_total=qtd_itens_total
)

if pdf_buffer is not None:
    st.download_button(
        label="📥 Exportar Relatório em PDF",
        data=pdf_buffer,
        file_name=f"relatorio_estoque_{pdv_selecionado}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
        mime="application/pdf",
        type="primary"
    )